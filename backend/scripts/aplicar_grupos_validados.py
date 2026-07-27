# Aplica a la base una reasignacion de grupos de movimiento decidida a mano en
# un excel de revision.
#
# Para que sirve: los gastos anteriores a cierta fecha llegaron de la migracion
# del sistema anterior SIN grupo, porque esa columna todavia no existia. Eso
# deja un hueco en cualquier analisis por grupo (en Power BI aparecen todos
# juntos en el bucket "(en blanco)"). Reetiquetarlos no se puede automatizar
# del todo: hay que mirarlos uno por uno. De paso, el catalogo de grupos
# acumulo etiquetas que quedaron sin sentido y conviene fusionar o borrar.
#
# El flujo es:
#   1. Se exporta un excel con una fila por DESCRIPCION distinta (no por
#      movimiento: muchos movimientos comparten descripcion, asi que revisar
#      por descripcion es un orden de magnitud menos trabajo).
#   2. Una persona completa la columna "GRUPO VALIDADO".
#   3. Este script lee ese excel y lo aplica, en una sola transaccion.
#
# El excel NO vive en el repo: lleva descripciones y montos reales, asi que va
# en datos_reales/ (gitignoreado), igual que el excel de la migracion original.
# Su estructura es generica a proposito -el nombre de cada categoria viaja como
# dato en una columna, nunca en el nombre de una hoja- para que este script no
# tenga que conocer ninguna categoria del negocio. Las hojas que espera estan
# documentadas en la hoja "Instrucciones" del propio excel.
#
# Es idempotente: volver a correrlo no cambia nada, porque cada paso busca los
# movimientos que TODAVIA estan en el estado de origen.
#
# Uso:  python backend/scripts/aplicar_grupos_validados.py [--dry-run] [--excel RUTA]
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict

import openpyxl
import psycopg2

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # backend/
DATOS_REALES = os.path.join(os.path.dirname(BASE), "datos_reales")

DRY_RUN = "--dry-run" in sys.argv
if "--excel" in sys.argv:
    EXCEL = sys.argv[sys.argv.index("--excel") + 1]
else:
    EXCEL = os.path.join(DATOS_REALES, "grupos_a_validar.xlsx")


def conexion():
    envvars = {}
    with open(os.path.join(BASE, ".env"), encoding="utf-8") as f:
        for linea in f:
            m = re.match(r"^\s*([^#=]+?)\s*=\s*(.*)\s*$", linea)
            if m:
                envvars[m.group(1)] = m.group(2)
    return psycopg2.connect(
        host=envvars["DB_HOST"], port=envvars["DB_PORT"],
        user=envvars["DB_USER"], password=envvars["DB_PASSWORD"],
        dbname=envvars["DB_NAME"],
    )


def norm(s):
    """Normaliza una descripcion para emparejar excel contra base. Tiene que
    ser IDENTICA a la que se uso al generar el excel, o no se encuentra nada."""
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s.upper().strip())
    return re.sub(r"\s+", " ", s).strip()


def columna(hoja, nombre):
    cab = [c.value for c in hoja[1]]
    if nombre not in cab:
        raise SystemExit(f"La hoja '{hoja.title}' no tiene la columna '{nombre}'")
    return cab.index(nombre)


# ------------------------------------------------------------------ leer excel
if not os.path.exists(EXCEL):
    raise SystemExit(
        f"No existe el excel de validacion: {EXCEL}\n"
        "Se espera en datos_reales/ o se pasa con --excel RUTA."
    )

wb = openpyxl.load_workbook(EXCEL, data_only=True)

h = wb["Config"]
i_clave, i_valor = columna(h, "CLAVE"), columna(h, "VALOR")
config = {f[i_clave]: f[i_valor] for f in h.iter_rows(min_row=2, values_only=True) if f[i_clave]}
corte = config.get("corte_sin_grupo")
if corte is None:
    raise SystemExit("Falta 'corte_sin_grupo' en la hoja Config del excel.")
corte = str(corte)[:10]

h = wb["Grupos"]
i_g, i_a = columna(h, "GRUPO"), columna(h, "ACCION")
acciones = {}
for f in h.iter_rows(min_row=2, values_only=True):
    if f[i_g]:
        acciones[str(f[i_g]).strip()] = str(f[i_a] or "mantener").strip().lower()

# Reasignaciones por descripcion, agrupadas por grupo de origen.
# Clave "" = movimientos que hoy no tienen grupo.
h = wb["Por descripcion"]
i_act, i_desc, i_val = (columna(h, "GRUPO ACTUAL"), columna(h, "DESCRIPCION"),
                        columna(h, "GRUPO VALIDADO"))
por_origen = defaultdict(dict)
for f in h.iter_rows(min_row=2, values_only=True):
    if f[i_val]:
        por_origen[str(f[i_act] or "").strip()][norm(f[i_desc])] = str(f[i_val]).strip()

h = wb["Por id"]
i_id, i_val = columna(h, "ID MOV"), columna(h, "GRUPO VALIDADO")
mapa_id = {}
for f in h.iter_rows(min_row=2, values_only=True):
    if f[i_val] and f[i_id] is not None:
        mapa_id[int(f[i_id])] = str(f[i_val]).strip()

print(f"Excel: {EXCEL}")
print(f"  corte para movimientos sin grupo: {corte}")
print(f"  grupos declarados: {len(acciones)} "
      f"(crear {sum(1 for a in acciones.values() if a == 'crear')}, "
      f"eliminar {sum(1 for a in acciones.values() if a == 'eliminar')})")
for origen, mapa in por_origen.items():
    print(f"  descripciones a reasignar desde {origen or '(sin grupo)'}: {len(mapa)}")
print(f"  movimientos por id: {len(mapa_id)}")

con = conexion()
cur = con.cursor()
cambios = Counter()
sin_encontrar = defaultdict(int)

# ------------------------------------------------------------- crear grupos
for grupo, accion in acciones.items():
    if accion == "crear":
        cur.execute(
            'INSERT INTO "Grupo_Movimiento" ("Nombre_Grupo_Movimiento") VALUES (%s) '
            'ON CONFLICT ("Nombre_Grupo_Movimiento") DO NOTHING',
            (grupo,),
        )
        if cur.rowcount:
            print(f"  + grupo creado: {grupo}")

cur.execute('SELECT "Nombre_Grupo_Movimiento", "Id_Grupo_Movimiento" FROM "Grupo_Movimiento"')
id_por_grupo = {n: i for n, i in cur.fetchall()}

pedidos = set()
for mapa in por_origen.values():
    pedidos |= set(mapa.values())
pedidos |= set(mapa_id.values())
faltantes = sorted(g for g in pedidos if g not in id_por_grupo)
if faltantes:
    raise SystemExit(
        "El excel manda movimientos a grupos que no existen y que tampoco estan "
        "marcados como 'crear' en la hoja Grupos: " + ", ".join(faltantes)
    )


def asignar(ids, grupo):
    if not ids:
        return
    # El IS DISTINCT FROM hace que el informe cuente cambios de verdad: si se
    # vuelve a correr sobre una base ya reasignada, el total da cero.
    cur.execute(
        'UPDATE "Movimiento" SET "Id_Grupo_Movimiento" = %s '
        'WHERE "Id_Movimiento" = ANY(%s) AND "Id_Grupo_Movimiento" IS DISTINCT FROM %s',
        (id_por_grupo[grupo], list(ids), id_por_grupo[grupo]),
    )
    cambios[grupo] += cur.rowcount


# --------------------------------------------------- 1) reasignar por descripcion
# Sin grupo de origen: el hueco de la migracion. Se acota por fecha y se
# excluyen las salidas que no son gastos (compras, pagos y servicios tienen su
# propia tabla y su propio vinculo; ver decision "categorizar sin adivinar").
SQL_SIN_GRUPO = '''
    SELECT m."Id_Movimiento", m."Descripcion_Movimiento"
    FROM "Movimiento" m
    WHERE m."Tipo_Movimiento" = 'SALIDA'
      AND m."Id_Grupo_Movimiento" IS NULL
      AND m."Fecha_Movimiento" < %s
      AND m."Id_Movimiento" NOT IN (
            SELECT "Id_Movimiento" FROM "Compra" WHERE "Id_Movimiento" IS NOT NULL)
      AND m."Id_Movimiento" NOT IN (
            SELECT "Id_Movimiento" FROM "Pago_Trabajador" WHERE "Id_Movimiento" IS NOT NULL)
      AND m."Id_Movimiento" NOT IN (
            SELECT "Id_Movimiento" FROM "Gasto_Extra_Mes" WHERE "Id_Movimiento" IS NOT NULL)'''

SQL_DESDE_GRUPO = '''
    SELECT m."Id_Movimiento", m."Descripcion_Movimiento"
    FROM "Movimiento" m
    JOIN "Grupo_Movimiento" g ON g."Id_Grupo_Movimiento" = m."Id_Grupo_Movimiento"
    WHERE g."Nombre_Grupo_Movimiento" = %s'''

for origen, mapa in por_origen.items():
    if origen:
        cur.execute(SQL_DESDE_GRUPO, (origen,))
    else:
        cur.execute(SQL_SIN_GRUPO, (corte,))
    destino = defaultdict(list)
    for id_mov, desc in cur.fetchall():
        grupo = mapa.get(norm(desc))
        if grupo:
            destino[grupo].append(id_mov)
        else:
            sin_encontrar[f"{origen or '(sin grupo)'}: {norm(desc)}"] += 1
    for grupo, ids in destino.items():
        asignar(ids, grupo)

# --------------------------------------------------------- 2) reasignar por id
destino = defaultdict(list)
for id_mov, grupo in mapa_id.items():
    destino[grupo].append(id_mov)
for grupo, ids in destino.items():
    asignar(ids, grupo)

# ------------------------------------------------------------ 3) borrar grupos
# Solo los que quedaron sin ningun movimiento. Si alguno todavia tiene, no se
# toca y se avisa, en vez de reventar con un error de clave foranea.
borrados, no_borrados = [], []
for grupo, accion in acciones.items():
    if accion != "eliminar" or grupo not in id_por_grupo:
        continue
    cur.execute('SELECT COUNT(*) FROM "Movimiento" WHERE "Id_Grupo_Movimiento" = %s',
                (id_por_grupo[grupo],))
    n = cur.fetchone()[0]
    if n:
        no_borrados.append((grupo, n))
    else:
        cur.execute('DELETE FROM "Grupo_Movimiento" WHERE "Id_Grupo_Movimiento" = %s',
                    (id_por_grupo[grupo],))
        borrados.append(grupo)

# ------------------------------------------------------------------- informe
print("\nMovimientos reasignados por grupo destino:")
for grupo, n in sorted(cambios.items(), key=lambda x: -x[1]):
    print(f"  {grupo:26s} {n:6d}")
print(f"  {'TOTAL':26s} {sum(cambios.values()):6d}")

if borrados:
    print("\nGrupos eliminados (quedaron vacios): " + ", ".join(sorted(borrados)))
for grupo, n in no_borrados:
    print(f"\nOJO: '{grupo}' no se borro: todavia tiene {n} movimientos.")
if sin_encontrar:
    total = sum(sin_encontrar.values())
    print(f"\nOJO: {total} movimientos ({len(sin_encontrar)} descripciones) sin "
          f"correspondencia en el excel; quedan como estaban. Primeras:")
    for d, n in sorted(sin_encontrar.items(), key=lambda x: -x[1])[:10]:
        print(f"    {n:4d}x  {d[:70]}")

if DRY_RUN:
    con.rollback()
    print("\n--dry-run: ROLLBACK, no se guardo nada.")
else:
    con.commit()
    print("\nCambios guardados.")
con.close()
